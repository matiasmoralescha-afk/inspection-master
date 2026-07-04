"""
Sunkist Container Status parser.

ops@elitequalityassurance.com sends an .xlsx attachment twice a day (subject
"Sunkist Container Status (YYYY-MM-DD_HHMMSS)") consolidating every Sunkist
container currently in the pipeline — confirmed 07/2026 as the only complete
source: of the 15 vessels seen in one report, only 3 also get an ad-hoc
"<VESSEL> - SUNKIST GLOBAL - <COMMODITY> - ETA ..." announcement thread from
721 Logistics. The xlsx is required; the vessel threads are not parsed.

Sheet layout (row 2 is the header, data starts at row 3):
  Entry# | Vessel | ETA | Origin | Commodity | Port/Terminal | Shipper |
  Container | Bill of Lading | Treatment | Status | USDA CT | ISF Filed |
  Trucker | Warehouse | Loc. | Customer

Each container is a "master" row (Container + Vessel both present) followed
by zero or more event-log rows starting with "-->" in column A, e.g.:
  --> | 7/2, 2:45 PM (EST) by ... | | | MNBU4052480 - Gate In | ...
"Gate In" is the warehouse-arrival signal — the equivalent of the COMMENTS
"DELIVERED" phrase the ocean-report parser looks for via Claude Haiku, but
explicit and structured here.

Vessel-name section-header rows (e.g. "POLAR BRASIL into PHIL", col A only,
no Container) are skipped.
"""
import io
import logging
import re
from datetime import datetime
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

_SUBJECT_RE = re.compile(r'^Sunkist Container Status\b', re.IGNORECASE)

# Column indices (0-based) matching the header row described above.
_COL_ENTRY, _COL_VESSEL, _COL_ETA, _COL_ORIGIN, _COL_COMMODITY = 0, 1, 2, 3, 4
_COL_CONTAINER, _COL_BL, _COL_WAREHOUSE = 7, 8, 14

_GATE_IN_RE = re.compile(r'gate\s*in', re.IGNORECASE)
_EVENT_DATE_RE = re.compile(r'(\d{1,2})/(\d{1,2})')


def is_sunkist_status(subject: str) -> bool:
    """True for 'Sunkist Container Status (...)' subjects."""
    return bool(_SUBJECT_RE.match(subject or ''))


def _cell_date(value) -> Optional[str]:
    """openpyxl returns ETA cells as datetime objects (data_only=True)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, 'isoformat'):  # date
        return value.isoformat()
    return None


def _event_date(raw: Optional[str]) -> Optional[str]:
    """'7/2, 2:45 PM (EST)  by amcclellan' -> 'YYYY-MM-DD' (assumes current year —
    events are always near-term relative to the report they appear in)."""
    if not raw:
        return None
    m = _EVENT_DATE_RE.search(raw)
    if not m:
        return None
    month, day = m.groups()
    year = datetime.now().year
    try:
        return f'{year:04d}-{int(month):02d}-{int(day):02d}'
    except ValueError:
        return None


def parse_xlsx(data: bytes) -> list[dict]:
    """
    Parse the Container Status .xlsx into one dict per container:
        {entry_no, vessel, eta_fecha, origin, commodity_raw, unit_id,
         bill_of_lading, warehouse, warehouse_arrival_confirmed,
         warehouse_arrival_at, latest_event}
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    records: list[dict] = []
    current: Optional[dict] = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        col0 = row[0]

        if isinstance(col0, str) and col0.strip() == '-->':
            if current is None:
                continue
            event_cell = row[_COL_COMMODITY] or ''
            event_type = event_cell.split(' - ', 1)[1].strip() if ' - ' in event_cell else event_cell.strip()
            event_date = _event_date(row[1])
            current['events'].append({'type': event_type, 'date': event_date})
            continue

        container = row[_COL_CONTAINER]
        vessel = row[_COL_VESSEL]
        if container and vessel:
            if current is not None:
                records.append(current)
            current = {
                'entry_no': str(col0).strip() if col0 else None,
                'vessel': str(vessel).strip(),
                'eta_fecha': _cell_date(row[_COL_ETA]),
                'origin': (row[_COL_ORIGIN] or '').strip() or None,
                'commodity_raw': (row[_COL_COMMODITY] or '').strip() or None,
                'unit_id': str(container).strip(),
                'bill_of_lading': (row[_COL_BL] or '').strip() or None,
                'warehouse': (row[_COL_WAREHOUSE] or '').strip() or None,
                'events': [],
            }
            continue
        # else: section-header row ("<VESSEL> into <PORT>") or blank — skip

    if current is not None:
        records.append(current)

    for rec in records:
        # Events are appended in sheet order, which lists the most recent
        # event first (observed: a 7/2 Gate In appears before a 6/30 Picked
        # Up, before a 6/29 Discharge) — so index 0 is the latest occurrence.
        gate_in_events = [e for e in rec['events'] if _GATE_IN_RE.search(e['type'])]
        rec['warehouse_arrival_confirmed'] = bool(gate_in_events)
        rec['warehouse_arrival_at'] = gate_in_events[0]['date'] if gate_in_events else None
        rec['latest_event'] = rec['events'][0]['type'] if rec['events'] else None
        del rec['events']

    if not records:
        logger.warning('Sunkist Container Status: no containers parsed from attachment')

    return records
